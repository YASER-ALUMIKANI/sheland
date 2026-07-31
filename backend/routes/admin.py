"""
CityLand Backend - Admin Management Routes
"""
import io
import logging
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..database import get_db
from .. import models, schemas, auth, analytics

logger = logging.getLogger("sheland.api")

router = APIRouter()


@router.post("/api/admin/users", response_model=schemas.UserResponse, status_code=201, include_in_schema=False)
def admin_create_user(
    request: Request,
    user_in: schemas.AdminUserCreate,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.require_roles(["admin", "super_admin"]))
):
    """Admin-only endpoint to create new users with any role (Admin, Sales Manager, Seller, Customer)."""
    ADMIN_MANAGED_ROLES = {"admin", "super_admin", "sales_manager", "seller", "customer"}
    target_role = user_in.role.lower().strip()
    if target_role not in ADMIN_MANAGED_ROLES:
        raise HTTPException(status_code=400, detail="الدور المحدد غير صالح")

    if target_role == "super_admin" and current_admin.role != "super_admin":
        raise HTTPException(status_code=403, detail="فقط المدير الفائق (Super Admin) يمكنه إنشاء حسابات Super Admin")

    existing_user = db.query(models.User).filter(
        (models.User.email == user_in.email) | (models.User.phone == user_in.phone)
    ).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني أو رقم الهاتف مسجل بالفعل")

    new_user = models.User(
        name=user_in.name,
        email=user_in.email,
        phone=user_in.phone,
        password_hash=auth.hash_password(user_in.password),
        role=target_role
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    if target_role == "seller":
        seller = models.Seller(user_id=new_user.id, store_name=f"متجر {new_user.name}")
        db.add(seller)
        db.commit()

    audit = models.AuditLog(
        action="admin_create_user",
        target_user_id=new_user.id,
        performed_by=current_admin.id,
        new_value=target_role,
        ip_address=request.client.host if (request and request.client) else "unknown",
        details=f"Admin '{current_admin.email}' created user '{new_user.email}' with role '{target_role}'"
    )
    db.add(audit)
    db.commit()

    return new_user


@router.put("/api/admin/users/{user_id}/role", response_model=schemas.UserResponse, include_in_schema=False)
def admin_update_user_role(
    user_id: int,
    role_in: schemas.UserRoleUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.require_roles(["admin", "super_admin"]))
):
    """Admin-only endpoint to promote or update a user's role with full audit logging."""
    ADMIN_MANAGED_ROLES = {"admin", "super_admin", "sales_manager", "seller", "customer"}
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    new_role = role_in.role.lower().strip()
    if new_role not in ADMIN_MANAGED_ROLES:
        raise HTTPException(status_code=400, detail="الدور المحدد غير صالح")

    if (new_role == "super_admin" or target_user.role == "super_admin") and current_admin.role != "super_admin":
        raise HTTPException(status_code=403, detail="فقط المدير الفائق (Super Admin) يمكنه تعديل أدوار Super Admin")

    old_role = target_user.role
    target_user.role = new_role
    db.commit()
    db.refresh(target_user)

    client_ip = request.client.host if (request and request.client) else "unknown"
    logger.warning(
        f"🔐 Security Audit: Role changed for user_id={target_user.id} ({target_user.email}) "
        f"from '{old_role}' to '{new_role}' by admin={current_admin.id} ({current_admin.email}) from IP={client_ip}"
    )

    audit = models.AuditLog(
        action="role_change",
        target_user_id=target_user.id,
        performed_by=current_admin.id,
        old_value=old_role,
        new_value=new_role,
        ip_address=request.client.host if (request and request.client) else "unknown",
        details=f"Admin '{current_admin.email}' changed user '{target_user.email}' role from '{old_role}' to '{new_role}'"
    )
    db.add(audit)
    db.commit()

    return target_user


@router.get("/api/admin/audit-logs", include_in_schema=False)
def get_audit_logs(
    limit: int = 50,
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(auth.require_roles(["admin", "super_admin"]))
):
    """Retrieve security audit trail logs."""
    logs = db.query(models.AuditLog).order_by(models.AuditLog.id.desc()).limit(limit).all()
    return logs


@router.get("/api/admin/analytics")
def get_admin_analytics(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles(["admin", "sales_manager"]))
):
    return analytics.compute_ecommerce_analytics(db)


@router.get("/api/admin/inventory/alerts")
def get_inventory_alerts(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles(["admin", "sales_manager", "seller"]))
):
    """Returns low stock products (<= 5 items) and stagnant products (no sales for 7+ or 30+ days)."""
    return analytics.calculate_inventory_alerts(db)


@router.get("/api/admin/reports/sales-excel")
def export_sales_excel_report(
    period: str = Query("all"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles(["admin", "sales_manager"]))
):
    """Generate and return styled Excel spreadsheet (.xlsx) for sales reports."""
    query = db.query(models.Order).order_by(models.Order.id.desc())
    now = datetime.utcnow()
    if period == "weekly":
        start_date = now - timedelta(days=7)
        query = query.filter(models.Order.created_at >= start_date)
    elif period == "monthly":
        start_date = now - timedelta(days=30)
        query = query.filter(models.Order.created_at >= start_date)

    orders = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير المبيعات الشامل"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    period_title = "الأسبوعي" if period == "weekly" else ("الشهري" if period == "monthly" else "الشامل")
    title_cell.value = f"منصة شي لاند — تقرير المبيعات والتخفيضات الرسمي ({period_title})"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="8B2C7C", end_color="8B2C7C", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    headers = [
        "رقم الطلب", "اسم العميل", "رقم الجوال", "عنوان التوصيل",
        "طريقة الدفع", "كود الخصم", "قيمة الخصم (ر.ي)", "إجمالي الطلب الصافي (ر.ي)",
        "حالة الدفع", "حالة الطلب", "تاريخ الطلب"
    ]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="1E1B2E")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    total_sales = 0.0
    total_discounts = 0.0

    for row_idx, o in enumerate(orders, start=4):
        disc = o.discount_amount or 0.0
        tot = o.total_amount or 0.0
        total_sales += tot
        total_discounts += disc
        date_str = o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else ""

        row_data = [
            o.order_number,
            o.customer_name or "عميل شي لاند",
            o.phone or "",
            o.shipping_address or "",
            o.payment_method or "COD",
            o.coupon_code or "-",
            disc,
            tot,
            o.payment_status or "pending",
            o.status or "قيد المعالجة",
            date_str
        ]
        ws.append(row_data)

        for col_num in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", horizontal="center" if col_num not in [2,4] else "right")
            if col_num in [7, 8]:
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 22

    last_row = len(orders) + 4
    ws.merge_cells(f"A{last_row}:F{last_row}")
    tot_label_cell = ws.cell(row=last_row, column=1)
    tot_label_cell.value = "الإجمالي التراكمي:"
    tot_label_cell.font = Font(name="Calibri", size=11, bold=True, color="8B2C7C")
    tot_label_cell.alignment = Alignment(horizontal="left", vertical="center")

    disc_tot_cell = ws.cell(row=last_row, column=7)
    disc_tot_cell.value = total_discounts
    disc_tot_cell.font = Font(name="Calibri", size=11, bold=True, color="C53030")
    disc_tot_cell.number_format = '#,##0.00'

    sales_tot_cell = ws.cell(row=last_row, column=8)
    sales_tot_cell.value = total_sales
    sales_tot_cell.font = Font(name="Calibri", size=12, bold=True, color="10B981")
    sales_tot_cell.number_format = '#,##0.00'

    tot_fill = PatternFill(start_color="FAF5FF", end_color="FAF5FF", fill_type="solid")
    for col_num in range(1, 12):
        cell = ws.cell(row=last_row, column=col_num)
        cell.fill = tot_fill
        cell.border = thin_border

    col_widths = {'A': 16, 'B': 22, 'C': 16, 'D': 30, 'E': 14, 'F': 14, 'G': 18, 'H': 22, 'I': 16, 'J': 16, 'K': 18}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Sheland_Sales_Report_{period}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
