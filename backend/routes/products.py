"""
CityLand Backend - Product & Category Routes
"""
import uuid
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func

from ..database import get_db
from .. import models, schemas, auth, cache
from backend.main import limiter, EXCEL_HEADERS, validate_and_save_image

router = APIRouter()


# --- Category Endpoints ---

@router.get("/api/categories", response_model=List[schemas.CategoryResponse])
def get_categories(db: Session = Depends(get_db)):
    cached = cache.get_cache("cache:categories")
    if cached is not None:
        return cached
    categories = db.query(models.Category).all()
    result = [schemas.CategoryResponse.from_orm(c) for c in categories]
    cache.set_cache("cache:categories", [c.dict() for c in result], expire_seconds=600)
    return result


# --- Product Endpoints ---

@router.get("/api/products", response_model=List[schemas.ProductResponse])
def get_products(
    category_id: Optional[int] = None,
    search: Optional[str] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    min_rating: Optional[float] = None,
    free_shipping: Optional[bool] = None,
    cod_available: Optional[bool] = None,
    sort_by: Optional[str] = "relevance",
    db: Session = Depends(get_db)
):
    cache_key = f"cache:products:{category_id or 'all'}:{sort_by}" if not search and min_price is None and max_price is None else None
    if cache_key:
        cached = cache.get_cache(cache_key)
        if cached is not None:
            return cached

    query = db.query(models.Product)

    if category_id:
        query = query.filter(models.Product.category_id == category_id)
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            (models.Product.title_ar.like(search_pattern)) |
            (models.Product.title_en.like(search_pattern)) |
            (models.Product.description.like(search_pattern))
        )
    if min_price is not None:
        query = query.filter(models.Product.price >= min_price)
    if max_price is not None:
        query = query.filter(models.Product.price <= max_price)
    if min_rating is not None:
        query = query.filter(models.Product.rating >= min_rating)
    if free_shipping:
        query = query.filter(models.Product.free_shipping == True)
    if cod_available:
        query = query.filter(models.Product.cod_available == True)

    if sort_by == "price_asc":
        query = query.order_by(models.Product.price.asc())
    elif sort_by == "price_desc":
        query = query.order_by(models.Product.price.desc())
    elif sort_by == "rating":
        query = query.order_by(models.Product.rating.desc())
    elif sort_by == "newest":
        query = query.order_by(models.Product.created_at.desc())
    else:
        query = query.order_by(models.Product.is_featured.desc(), models.Product.id.desc())

    products = query.all()
    results = [schemas.ProductResponse.from_orm_with_stock(p) for p in products]

    if cache_key:
        cache.set_cache(cache_key, [r.dict() for r in results], expire_seconds=300)

    return results


# --- Excel Bulk Import & Template Endpoints ---

@router.get("/api/products/excel-template")
def download_excel_template():
    """Generates an official styled Excel template workbook for bulk product upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قالب استيراد المنتجات"
    ws.views.sheetView[0].rightToLeft = True

    header_fill = PatternFill(start_color="5C2C77", end_color="5C2C77", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Segoe UI", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (field_key, header_title) in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    sample_data = [
        [
            "فستان سهرة أنيق ومميز", "Elegant Evening Dress", "أزياء نسائية",
            12500, 18000, 8500, 25, "DRS-LUX-01",
            "https://images.unsplash.com/photo-1566174053879-31528523f8ae",
            "فستان سهرة فاخر من القماش المخملي العالي الجودة", "أسود", "L", "نعم", "نعم"
        ],
        [
            "قميص رجالي كلاسيكي قطن", "Classic Men Cotton Shirt", "أزياء رجالية",
            6800, 9500, 4200, 40, "SHRT-M-02",
            "https://images.unsplash.com/photo-1602810318383-e386cc2a3ccf",
            "قميص قطن 100% مناسب للمناسبات والعمل", "أزرق", "XL", "نعم", "نعم"
        ]
    ]

    for row_idx, row_values in enumerate(sample_data, start=2):
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font

    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24

    ws.row_dimensions[1].height = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    headers = {"Content-Disposition": "attachment; filename=sheland_products_template.xlsx"}
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.post("/api/products/import-excel")
@limiter.limit("20/minute")
def import_products_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles(["seller", "admin", "super_admin"]))
):
    """Parses uploaded Excel (.xlsx) or CSV (.csv) file and imports products in bulk."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="يرجى رفع ملف صحيح")

    ext = file.filename.split(".")[-1].lower()
    if ext not in ["xlsx", "xls", "csv"]:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم، يرجى اختيار ملف Excel (.xlsx) أو CSV (.csv)")

    contents = file.file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="حجم الملف كبير جداً (الحد الأقصى 10 ميجابايت)")

    seller_id = None
    if current_user.role == "seller":
        seller = db.query(models.Seller).filter(models.Seller.user_id == current_user.id).first()
        if not seller:
            seller = models.Seller(user_id=current_user.id, store_name=f"متجر {current_user.name}")
            db.add(seller)
            db.commit()
            db.refresh(seller)
        seller_id = seller.id
    else:
        seller = db.query(models.Seller).first()
        seller_id = seller.id if seller else 1

    rows_data = []
    if ext in ["xlsx", "xls"]:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows or len(all_rows) < 2:
            raise HTTPException(status_code=400, detail="ملف الإكسل فارغ أو لا يحتوي على صفوف بيانات")
        
        headers_row = [str(cell or "").strip().lower() for cell in all_rows[0]]
        for r in all_rows[1:]:
            if any(r):
                row_dict = {headers_row[i]: r[i] for i in range(min(len(headers_row), len(r)))}
                rows_data.append(row_dict)
    else:
        text = contents.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            clean_row = {str(k or "").strip().lower(): v for k, v in row.items()}
            rows_data.append(clean_row)

    if not rows_data:
        raise HTTPException(status_code=400, detail="لم يتم العثور على بيانات منتجات قابلة للاستيراد بداخل الملف")

    def get_val(row_map, *keys):
        for k in keys:
            for map_key, val in row_map.items():
                if k in map_key and val is not None:
                    return val
        return None

    imported_count = 0
    failed_count = 0
    errors = []

    for idx, row in enumerate(rows_data, start=2):
        title_ar = get_val(row, "title_ar", "اسم المنتج", "عربي", "اسم")
        price_val = get_val(row, "price", "سعر البيع", "السعر")
        category_val = get_val(row, "category", "القسم", "التصنيف")

        if not title_ar or price_val is None:
            failed_count += 1
            errors.append(f"الصف {idx}: تم التجاوز لعدم وجود اسم المنتج أو السعر")
            continue

        try:
            price = float(price_val)
        except ValueError:
            failed_count += 1
            errors.append(f"الصف {idx}: السعر غير صحيح ({price_val})")
            continue

        title_ar = str(title_ar).strip()
        title_en = str(get_val(row, "title_en", "إنجليزي", "english") or title_ar).strip()

        category = None
        if category_val:
            cat_str = str(category_val).strip()
            category = db.query(models.Category).filter(
                (models.Category.name_ar == cat_str) |
                (models.Category.name_en == cat_str) |
                (models.Category.slug == cat_str.lower())
            ).first()

        if not category:
            category = db.query(models.Category).first()
            if not category:
                category = models.Category(name_ar="عام", name_en="General", slug="general")
                db.add(category)
                db.commit()
                db.refresh(category)

        compare_at = get_val(row, "compare_at_price", "قبل الخصم", "compare")
        cost_p = get_val(row, "cost_price", "التكلفة", "cost")
        stock_val = get_val(row, "stock", "المخزون", "الكمية")
        img_url = get_val(row, "image_url", "صورة", "رابط") or "https://images.unsplash.com/photo-1523275335684-37898b6baf30"
        desc = get_val(row, "description", "وصف", "تفاصيل")
        sku_val = get_val(row, "sku", "رمز", "كود")
        color_val = get_val(row, "color", "اللون")
        size_val = get_val(row, "size", "المقاس")
        free_shipping = str(get_val(row, "free_shipping", "شحن") or "نعم").strip().lower() not in ["لا", "false", "0", "no"]
        cod_avail = str(get_val(row, "cod_available", "دفع") or "نعم").strip().lower() not in ["لا", "false", "0", "no"]

        try:
            compare_price = float(compare_at) if compare_at is not None and str(compare_at).strip() != "" else None
        except ValueError:
            compare_price = None

        try:
            cost_price = float(cost_p) if cost_p is not None and str(cost_p).strip() != "" else 0.0
        except ValueError:
            cost_price = 0.0

        try:
            stock = int(float(stock_val)) if stock_val is not None and str(stock_val).strip() != "" else 10
        except ValueError:
            stock = 10

        sku_clean = str(sku_val).strip() if sku_val else None
        existing_product = None

        if sku_clean:
            v_match = db.query(models.ProductVariant).filter(models.ProductVariant.sku == sku_clean).first()
            if v_match:
                existing_product = db.query(models.Product).filter(models.Product.id == v_match.product_id).first()

        if not existing_product:
            existing_product = db.query(models.Product).filter(
                (models.Product.seller_id == seller_id) &
                (func.lower(models.Product.title_ar) == title_ar.lower())
            ).first()

        if existing_product:
            existing_product.price = price
            if compare_price is not None:
                existing_product.compare_at_price = compare_price
            if cost_price > 0:
                existing_product.cost_price = cost_price
            if img_url:
                existing_product.image_url = str(img_url).strip()
            existing_product.category_id = category.id
            existing_product.free_shipping = free_shipping
            existing_product.cod_available = cod_avail
            if desc:
                existing_product.description = str(desc)

            variant = None
            if sku_clean:
                variant = db.query(models.ProductVariant).filter(
                    (models.ProductVariant.product_id == existing_product.id) &
                    (models.ProductVariant.sku == sku_clean)
                ).first()
            if not variant:
                variant = db.query(models.ProductVariant).filter(models.ProductVariant.product_id == existing_product.id).first()

            if variant:
                variant.stock = (variant.stock or 0) + stock
                if color_val:
                    variant.color = str(color_val).strip()
                if size_val:
                    variant.size = str(size_val).strip()
                if sku_clean:
                    variant.sku = sku_clean
            else:
                variant = models.ProductVariant(
                    product_id=existing_product.id,
                    sku=sku_clean or f"SKU-{existing_product.id}",
                    color=str(color_val).strip() if color_val else None,
                    size=str(size_val).strip() if size_val else None,
                    stock=stock
                )
                db.add(variant)

            db.commit()
            imported_count += 1
        else:
            slug = f"seller-{seller_id}-{uuid.uuid4().hex[:6]}-{title_ar.replace(' ', '-')}"
            db_product = models.Product(
                seller_id=seller_id,
                category_id=category.id,
                title_ar=title_ar,
                title_en=title_en,
                slug=slug,
                description=str(desc) if desc else None,
                price=price,
                compare_at_price=compare_price,
                cost_price=cost_price,
                image_url=str(img_url).strip(),
                free_shipping=free_shipping,
                cod_available=cod_avail,
                is_featured=False
            )
            db.add(db_product)
            db.commit()
            db.refresh(db_product)

            db_variant = models.ProductVariant(
                product_id=db_product.id,
                sku=sku_clean or f"SKU-{db_product.id}",
                color=str(color_val).strip() if color_val else None,
                size=str(size_val).strip() if size_val else None,
                stock=stock
            )
            db.add(db_variant)
            db.commit()

            imported_count += 1

    cache.clear_cache_by_prefix("cache:products")
    return {
        "status": "success",
        "message": f"تم معالجة ودمج {imported_count} منتج بنجاح في المنصة!",
        "imported_count": imported_count,
        "failed_count": failed_count,
        "errors": errors
    }


@router.get("/api/products/{product_id}", response_model=schemas.ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return schemas.ProductResponse.from_orm_with_stock(product)


@router.post("/api/products", response_model=schemas.ProductResponse)
def create_product(
    product_in: schemas.ProductCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles(["seller"]))
):
    db_product = models.Product(
        seller_id=product_in.seller_id,
        category_id=product_in.category_id,
        title_ar=product_in.title_ar,
        title_en=product_in.title_en,
        slug=product_in.slug,
        description=product_in.description,
        price=product_in.price,
        compare_at_price=product_in.compare_at_price,
        cost_price=product_in.cost_price if product_in.cost_price is not None else round(product_in.price * 0.6),
        currency=product_in.currency,
        image_url=product_in.image_url,
        free_shipping=product_in.free_shipping,
        cod_available=product_in.cod_available,
    )
    db.add(db_product)
    db.commit()
    db.refresh(db_product)

    stock_val = getattr(product_in, 'stock', 10)
    if stock_val is None:
        stock_val = 10

    if product_in.variants and len(product_in.variants) > 0:
        for v in product_in.variants:
            db_variant = models.ProductVariant(
                product_id=db_product.id,
                sku=v.sku,
                color=v.color,
                size=v.size,
                stock=v.stock,
                price_override=v.price_override
            )
            db.add(db_variant)
    else:
        db_variant = models.ProductVariant(
            product_id=db_product.id,
            sku=f"SKU-{db_product.id}",
            stock=stock_val
        )
        db.add(db_variant)

    db.commit()
    db.refresh(db_product)
    cache.clear_cache_by_prefix("cache:")
    return schemas.ProductResponse.from_orm_with_stock(db_product)


@router.put("/api/products/{product_id}", response_model=schemas.ProductResponse)
def update_product(
    product_id: int,
    product_in: schemas.ProductCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles(["seller"]))
):
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")

    if current_user.role not in ["admin", "super_admin"]:
        seller = db.query(models.Seller).filter(models.Seller.user_id == current_user.id).first()
        if not seller or db_product.seller_id != seller.id:
            raise HTTPException(status_code=403, detail="غير مصرح: ليس لديك صلاحية تعديل هذا المنتج")

    db_product.title_ar = product_in.title_ar

    db_product.title_en = product_in.title_en
    db_product.category_id = product_in.category_id
    db_product.price = product_in.price
    if getattr(product_in, 'cost_price', None) is not None:
        db_product.cost_price = product_in.cost_price
    db_product.compare_at_price = product_in.compare_at_price
    db_product.image_url = product_in.image_url
    db_product.description = product_in.description
    db_product.free_shipping = product_in.free_shipping
    db_product.cod_available = product_in.cod_available

    p_stock = getattr(product_in, 'stock', None)
    if p_stock is not None:
        variants = db.query(models.ProductVariant).filter(models.ProductVariant.product_id == product_id).all()
        if variants:
            variants[0].stock = p_stock
            for v in variants[1:]:
                v.stock = 0
        else:
            db.add(models.ProductVariant(product_id=product_id, sku=f"SKU-{product_id}", stock=p_stock))

    db.commit()
    db.refresh(db_product)
    cache.clear_cache_by_prefix("cache:")
    return schemas.ProductResponse.from_orm_with_stock(db_product)


@router.delete("/api/products/{product_id}")
def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles(["admin"]))
):
    db_product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(db_product)
    db.commit()
    cache.clear_cache_by_prefix("cache:")
    return {"status": "success", "message": f"Product {product_id} deleted"}


@router.post("/api/upload")
@limiter.limit("20/minute")
async def upload_image_file(
    request: Request,
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_roles(["admin", "super_admin", "sales_manager", "seller"]))
):
    """Secure endpoint for sellers & admins to upload product photos with rate limiting & strict validation."""
    contents = await file.read()
    url = validate_and_save_image(contents, prefix="prod")
    return {"url": url}
