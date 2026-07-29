import io
import uuid
import openpyxl
from fastapi.testclient import TestClient
from backend.main import app
from tests.conftest import TestingSessionLocal
from backend import models, auth

client = TestClient(app)

def test_download_excel_template():
    """Test GET /api/products/excel-template returns a valid openpyxl spreadsheet."""
    response = client.get("/api/products/excel-template")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Load and verify Excel sheet
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    assert ws.title == "قالب استيراد المنتجات"
    
    # Check headers
    first_row = list(ws.iter_rows(values_only=True))[0]
    assert "اسم المنتج (عربي) *" in first_row
    assert "سعر التكلفة على البائع (ر.ي)" in first_row

def test_import_products_excel():
    """Test POST /api/products/import-excel uploads and creates products in bulk."""
    db = TestingSessionLocal()
    unique_email = f"excel_seller_{uuid.uuid4().hex[:6]}@sheland.com"
    seller_user = models.User(
        name="تاجر تجربة Excel",
        email=unique_email,
        phone=f"077{uuid.uuid4().hex[:6]}",
        password_hash=auth.hash_password("seller123"),
        role="seller"
    )
    db.add(seller_user)
    db.commit()
    db.refresh(seller_user)

    seller_profile = models.Seller(user_id=seller_user.id, store_name="متجر تجربة الإكسل")
    db.add(seller_profile)
    db.commit()

    token = auth.create_access_token({"sub": str(seller_user.id), "role": seller_user.role})
    db.close()
    headers = {"Authorization": f"Bearer {token}"}

    # 2. Build in-memory Excel workbook with test product data
    wb = openpyxl.Workbook()
    ws = wb.active
    headers_list = [
        "title_ar", "title_en", "category", "price", "compare_at_price",
        "cost_price", "stock", "sku", "image_url", "description", "color", "size"
    ]
    ws.append(headers_list)

    row1 = [
        "ساعة يد رجالية استيل", "Steel Men Watch", "إلكترونيات",
        15000, 22000, 9500, 15, "WTC-ST-01",
        "https://images.unsplash.com/photo-1523275335684-37898b6baf30",
        "ساعة رجالية سوداء فاخرة", "فضائي", "Standard"
    ]
    row2 = [
        "حقيبة ظهر مدرسية أنيقة", "School Backpack", "أزياء نسائية",
        8500, 12000, 5000, 30, "BAG-SCH-02",
        "https://images.unsplash.com/photo-1553062407-98eeb64c6a62",
        "حقيبة ظهر مريحة وعملية للطلاب", "رمادي", "Medium"
    ]
    ws.append(row1)
    ws.append(row2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # 3. Post to /api/products/import-excel
    files = {"file": ("test_import.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    res = client.post("/api/products/import-excel", headers=headers, files=files)

    assert res.status_code == 200
    data = res.json()
    assert data["status"] == "success"
    assert data["imported_count"] == 2

    # 4. Verify in DB
    check_db = TestingSessionLocal()
    prod1 = check_db.query(models.Product).filter(models.Product.title_ar == "ساعة يد رجالية استيل").first()
    assert prod1 is not None
    assert prod1.price == 15000
    assert prod1.cost_price == 9500
    assert prod1.compare_at_price == 22000

    # 5. Import SAME file again (Duplicate Test) -> Should merge stock (15 + 15 = 30) instead of duplicate row
    buffer.seek(0)
    files2 = {"file": ("test_import_dup.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    res2 = client.post("/api/products/import-excel", headers=headers, files=files2)
    assert res2.status_code == 200

    # Verify no duplicate product row was created and stock was merged
    prods = check_db.query(models.Product).filter(models.Product.title_ar == "ساعة يد رجالية استيل").all()
    assert len(prods) == 1
    assert prods[0].stock == 30  # 15 + 15 merged!
    check_db.close()
